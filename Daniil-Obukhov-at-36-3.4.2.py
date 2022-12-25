import csv
import os
import pathlib
from typing import List, Dict
import re
import doctest
import unittest
import numpy as np
import pandas as pd
import openpyxl
import requests
from matplotlib import pyplot as plt
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from jinja2 import Environment, FileSystemLoader
import pdfkit
from requests.adapters import HTTPAdapter
from urllib3 import Retry
from xlsx2html import xlsx2html
import time
from multiprocessing import Pool
import concurrent.futures


class ProcessValutes:
    def __init__(self, date, salary_currency):
        self.date = date
        self.salary_currency = salary_currency

    def get_currency_valute(self):
        if self.salary_currency == "RUR":
            return 1
        valutes = pd.read_csv("valutes.csv")
        valute = valutes.loc[valutes["date"] == self.date]
        if valute.__contains__(self.salary_currency):
            return float(valute[self.salary_currency])
        return 0


class Salary:
    """
    Класс для представления зарплат
    Attributes:
        salary_from (str): Нижняя граница оклада
        :type (str or int or float)
        salary_to: Верхняя граница оклада
        :type (str or int or float)
        salary_currency: Валюта оклада
        :type (str)
        published_at: Дата публикации
        :type (str)
    """
    def __init__(self, salary_from : str or int or float, salary_to : str or int or float, salary_currency : str, published_at : str):
        """
        @param salary_from: Нижняя граница оклада
        :type (str or int or float)
        @param salary_to: Верхняя граница оклада
        :type (str or int or float)
        @param salary_currency: Валюта оклада
        :type (str)
        @param published_at: Дата публикации
        :type (str)
        """
        self.salary_from = self.__check_void_value(salary_from)
        self.salary_to = self.__check_void_value(salary_to)
        self.salary_currency = salary_currency
        self.published_at = published_at
        self.month_year = f"{self.published_at[5:7]}/{self.published_at[:4]}"

    @staticmethod
    def __check_void_value(value: str or int or float) -> float:
        if type(value) == str and value == "":
            return 0
        return float(value)

    def get_average_salary(self):
        return round(((self.salary_from + self.salary_to) * ProcessValutes(self.month_year, self.salary_currency).get_currency_valute()) / 2, 4)


class Vacancy:
    """
    Класс для представления вакансий
    Attributes:
        name: Название
        :type (str)
        salary: Зарплата для данной вакансии
        :type (Salary)
        area_name: Местоположение
        :type (str)
        published_at: Дата публикации
        :type (str)
        year: Год публикации
        :type (str)
    """
    def __init__(self, vacancy: Dict[str, str]):
        """
        @param vacancy: Отдельная вакансия в виде словаря: атрибут - значение
        :type (Dict[str, str])
        >>> vac = {"name" :"Инженер", "salary_from" : 35000.0,"salary_to" : 45000.0, "salary_currency" : "RUR", "area_name" : "Moscow","published_at" :"2007-12-03T17:47:55+0300"}
        >>> vac = Vacancy(vac)
        >>> vac.area_name
        'Moscow'
        >>> vac.year
        '2007'
        """
        self.name = vacancy["name"]
        self.salary = Salary(salary_from=vacancy["salary_from"],
                             salary_to=vacancy["salary_to"],
                             salary_currency=vacancy["salary_currency"],
                             published_at=vacancy["published_at"])
        self.area_name = vacancy["area_name"]
        self.published_at = vacancy["published_at"]
        self.year = self.published_at[:4]

    def get_array_vacancy(self) -> List[str]:
        return [self.name, self.salary.get_average_salary(), self.area_name, self.published_at]


class SplitCsvFileByYear:
    """
    Класс для раделения набора вакансий по годам
    Attributes:
        file_name: Название файла
        :type (str)
        dir_name: Название папки, в которой хранятся итоговые csv-файлы
        :type (List[Vacancy])
        headlines: Названия загаловков
        :type (List[str])
        vacancies: Набор вакансий
        :type (List[List[str]])
    """
    def __init__(self, file_name : str, directory : str):
        """
        @param file_name: Название файла
        :type (str)
        @param file_name: Название папки, в которой хранятся итоговые csv-файлы
        :type (str)
        """
        self.file_name = file_name
        self.dir_name = directory
        self.headlines, self.vacancies = self.__csv_reader()
        self.__csv_process(self.headlines, self.vacancies)

    def __csv_reader(self) -> (List[str], List[List[str]]):
        """
        Читает из csv файла вакансии и возвращает в виде списка загаловков и набора вакансий
        @return: Список загаловков и набора вакансий
        :type (List[str], List[List[str]])
        """
        with open(self.file_name, encoding='utf-8-sig') as file:
            file_reader = csv.reader(file)
            lines = [row for row in file_reader]
        return lines[0], lines[1:]

    def __csv_process(self, headlines : List[str], vacancies : List[List[str]]) -> None:
        """
        Обрабатывает полученный набор вакансий и загаловков
        @param headlines: Названия загаловков
        :type (List[str])
        @param vacancies: Набор вакансий
        :type (List[List[str]])
        @return: None
        """
        cur_year = "0"
        self.first_vacancy = ""
        os.mkdir(self.dir_name)
        vacancies_cur_year = []
        for vacancy in vacancies:
            if (len(vacancy) == len(headlines)) and ((all([v != "" for v in vacancy])) or (vacancy[1] == "" and vacancy[2] != "") or (vacancy[1] != "" and vacancy[2] == "")):
                vacancy = [" ".join(re.sub("<.*?>", "", value).replace('\n', '; ').split()) for value in vacancy]
                if len(self.first_vacancy) == 0:
                    self.first_vacancy = vacancy
                vacancy_list = [v for v in vacancy]
                if vacancy[-1][:4] != cur_year:
                    if len(vacancies_cur_year) != 0:
                        self.__csv_writer(headlines, vacancies_cur_year, cur_year)
                        vacancies_cur_year.clear()
                    cur_year = vacancy[-1][:4]
                vacancies_cur_year.append(vacancy_list)
                self.last_vacancy = vacancy
        self.__csv_writer(headlines, vacancies_cur_year, cur_year)

    def __csv_writer(self, headlines : List[str], vacancies : List[List[str]], cur_year : str) -> None:
        """
        Записывает данные в csv-файл
        @param headlines: Названия загаловков
        :type (List[str])
        @param vacancies: Набор вакансий
        :type (List[List[str]])
        @param cur_year: Текущий год обработки
        :type (str)
        @return: None
        """
        name = os.path.splitext(self.file_name)
        vacancies = pd.DataFrame(vacancies, columns=headlines)
        vacancies.to_csv(f'{self.dir_name}/{name[0]}_{cur_year}.csv', index=False)


class DataSet:
    """
    Класс для представления набора вакансий
    Attributes:
        file_name: Название файла
        :type (str)
        vacancies_objects: писок вакансий
        :type (List[Vacancy])
    """
    def __init__(self, file_name : str):
        """
        @param file_name: Название файла
        :type (str)
        """
        self.file_name = file_name
        self.vacancies_objects = self.__csv_reader()

    def __csv_reader(self) -> (List[Vacancy]):
        """
        Читает из csv файла вакансии и возвращает в виде списка вакансий
        @return: Список вакансий
        :type (List[Vacancy])
        """
        with open(self.file_name, encoding='utf-8-sig') as file:
            file_reader = csv.reader(file)
            lines = [row for row in file_reader]
            headlines, vacancies = lines[0], lines[1:]
        return self.process_vacancies(headlines, vacancies)

    def process_vacancies(self, headlines : List[str], vacancies : List[List[str]]) -> (List[Vacancy]):
        """
        Отбирает правильно заполненные вакансии и конвертирует в класс Vacancy
        :param headlines: Названия заголовков
        :type (List[str])
        :param vacancies: Список из списокв вакансий
        :type (List[List[str]])
        :return: Правильно заполненные вакансии
        :type (List[Vacancy])
        """
        result = []
        for vacancy in vacancies:
            vacancy = [" ".join(re.sub("<.*?>", "", value).replace('\n', '; ').split()) for value in vacancy]
            result.append(Vacancy({x: y for x, y in zip([r for r in headlines], [v for v in vacancy])}))
        return result


class YearSalary:
    """
    Класс для представления параметра и связанной с ним зарплаты
    Attributes:
        param: Выбранный параметр соотносимый с определенной зарплатой
        :type (str)
        salary: Зарплата для определенной вакансии
        :type (Salary)
        currency_to_rub: Таблица перевода из определенной валюты в рубли
        :type (Dict[str, float])
        count_vacancies: Количество вакансий:
        :type (int)
    """
    def __init__(self, param : str, salary: Salary):
        """
        @param param: Выбранный параметр соотносимый с определенной зарплатой
        :type (str)
        @param salary: Зарплата для определенной вакансии
        :type (Salary)
        >>> YearSalary("year", Salary(1000, 2000, "USD")).salary
        90990.0
        """
        self.param = param
        self.salary = salary.get_average_salary()
        self.count_vacancies = 1

    def add_salary(self, new_salary : Salary):
        """
        Добавление зарплаты к выбранному параметру
        @param new_salary: Зарплата для добавления
        :type (Salary)
        @return: None
        >>> YearSalary("city", Salary(100, 200, "RUR")).add_salary(Salary(200, 300, "RUR")).salary
        400.0
        >>> YearSalary("city", Salary(100, 200, "RUR")).add_salary(Salary(200, 300, "RUR")).count_vacancies
        2
        """
        self.count_vacancies += 1
        self.salary = self.salary + new_salary.get_average_salary()
        return self


class Report:
    """
    Класс для создания Excel таблицы
    Attributes:
        profession: Профессия
        :type (str)
        years: Года представленных данных
        :type (str)
        average_salary: Набор средних зарплат по годам
        :type (List[int])
        average_salary_profession: Набор средних зарплат по годам для выбранной профессии
        :type (List[int])
        count_vacancies_by_year: Набор количеств вакансий за год
        :type (List[int])
        count_vacancies_by_year_prof: Набор количеств вакансий за год для выбранной професии
        :type (List[int])
        file_name: Название файла
        :type (str)
    """
    def __init__(self, profession : str, years: List[int], average_salary : List[int],
                 average_salary_profession : List[int], count_vacancies_by_year : List[int],
                 count_vacancies_by_year_prof : List[int], file_name : str):
        """
        @param profession: Профессия
        :type (str)
        @param years: Года представленных данных
        :type (List[int])
        @param average_salary: Набор средних зарплат по годам
        :type (List[int])
        @param average_salary_profession: Набор средних зарплат по годам для выбранной профессии
        :type (List[int])
        @param count_vacancies_by_year: Набор количеств вакансий за год
        :type (List[int])
        @param count_vacancies_by_year_prof: Набор количеств вакансий за год для выбранной професии
        :type (List[int])
        @param file_name: Название файла
        :type (str)
        """
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.profession = profession
        self.file_name = file_name

    def generate_excel(self) -> None:
        """
        Генерирует файл для полученных данных
        @return: None
        """
        if not isinstance(self.file_name, str):
            raise TypeError('')
        if os.path.basename(self.file_name).split('.')[1] != "xlsx":
            raise TypeError('')
        if os.path.exists(self.file_name):
            raise FileExistsError("")
        df = [[self.years[i], self.average_salary[i], self.average_salary_profession[i], self.count_vacancies_by_year[i], self.count_vacancies_by_year_prof[i]] for i in range(len(self.years))]
        df.insert(0, ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий", f"Количество вакансий - {self.profession}"])
        df = pd.DataFrame(df, columns=None)
        with pd.ExcelWriter(self.file_name) as writer:
            df.to_excel(writer, sheet_name='Статистика по годам', index=False, header=False)
        wb = openpyxl.load_workbook(self.file_name)
        worksheet1 = wb["Статистика по годам"]
        thin = Side(border_style="thin")
        self.__add_border_and_align(worksheet1, thin, len(self.years) + 2, ["A", "B", "C", "D", "E"])
        self.__make_max_column_width(worksheet1)
        wb.save(self.file_name)

    def __add_border_and_align(self, worksheet : Worksheet, side : Side, count_columns : int, rows : List[str]) -> None:
        """
        @param worksheet: Рабочий лист
        :type (Worksheet)
        @param side: Сторона печати
        :type (Side)
        @param count_columns: Количество столбцов
        :type (int)
        @param rows: Список выбранных названий строк
        :type (List[str])
        @return: None
        """
        for i in range(1, count_columns):
            for row in rows:
                if i == 1:
                    worksheet[row + str(i)].alignment = Alignment(horizontal='left')
                    worksheet[row + str(i)].font = Font(bold=True)
                if worksheet[row + str(i)].internal_value != None:
                    worksheet[row + str(i)].border = Border(top=side, bottom=side, left=side, right=side)

    def __make_max_column_width(self, worksheet : Worksheet) -> None:
        """
        @param worksheet: Рабочий лист
        :type (Worksheet)
        @return: None
        """
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value != None:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                else:
                    dims[cell.column] = len(str(cell.value))
        for col, value in dims.items():
            worksheet.column_dimensions[get_column_letter(col)].width = value + 2


class Graphic:
    """
    Класс для создания графиков с помощью библиотеки matpolib
    Attributes:
        profession: Профессия
        :type (str)
        years: Года представленных данных
        :type (str)
        average_salary: Набор средних зарплат по годам
        :type (List[int])
        average_salary_profession: Набор средних зарплат по годам для выбранной профессии
        :type (List[int])
        count_vacancies_by_year: Набор количеств вакансий за год
        :type (List[int])
        count_vacancies_by_year_prof: Набор количеств вакансий за год для выбранной професии
        :type (List[int])
        city_salary: Словарь с предмтавлением: город - зарплата
        :type (Dict[str, int])
        city_count_vacancies: Словарь с предмтавлением: город - количество вакансий
        :type (Dict[str, int])
        file_name: Название файла
        :type (str)
    """
    def __init__(self, profession: str, years: List[int], average_salary: List[int],
                 average_salary_profession: List[int], count_vacancies_by_year: List[int],
                 count_vacancies_by_year_prof: List[int], file_name : str):
        """
        @param profession: Выбранная профессия
        :type (str)
        @param years: Года представленных данных
        :type (List[int])
        @param average_salary: Набор средних зарплат по годам
        :type (List[int])
        @param average_salary_profession: Набор средних зарплат по годам для выбранной профессии
        :type (List[int])
        @param count_vacancies_by_year: Набор количеств вакансий за год
        :type (List[int])
        @param count_vacancies_by_year_prof: Набор количеств вакансий за год для выбранной професии
        :type (List[int])
        @param city_salary: Словарь с предмтавлением: город - зарплата
        :type (Dict[str, int])
        @param city_count_vacancies: Словарь с предмтавлением: город - количество вакансий
        :type (Dict[str, int])
        @param file_name: Название файла
        :type (str)
        """
        if not isinstance(file_name, str):
            raise TypeError('')
        if os.path.basename(file_name).split('.')[1] != "png":
            raise TypeError('')
        if os.path.exists(file_name):
            raise FileExistsError("")
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.profession = profession
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 8))
        self.__grouped_bar_graph(ax1, "Уровень зарплат по годам", self.average_salary, self.years,
                                 self.average_salary_profession, 'средняя з/п', f'з/п {self.profession}')
        self.__grouped_bar_graph(ax2, 'Количество вакансий по годам', self.count_vacancies_by_year, self.years,
                                 self.count_vacancies_by_year_prof, 'Количество вакансий',
                                 f'Количество вакансий {self.profession}')
        #self.__horizontal_graph(ax3)
        #self.__pie_graph(ax4)
        plt.tight_layout()
        #plt.show()
        fig.savefig(file_name)

    def __grouped_bar_graph(self, ax, title: str, values_x: List[int], values_y: List[int], values_x2: List[int],
                            label_x: str, label_x2: str) -> None:
        """
        Создание сгруппированной гистограммы
        @param ax: Местоположение графика
        :type (matplotlib.axes._subplots.AxesSubplot)
        @param title: Название графика
        :type (str)
        @param values_x: Значения 1 выборки по оси X
        :type (List[int])
        @param values_y: Значения  выборки по оси Y
        :type (List[int])
        @param values_x2: Значения 2 выборки по оси X
        :type (List[int])
        @param label_x: Название легенды для 1 выборки
        :type (str)
        @param label_x2: Название легенды для 2 выборки
        :type (str)
        @return: None
        """
        ax.grid(axis='y')
        x = np.arange(len(values_y))
        width = 0.4
        ax.bar(x - width / 2, values_x, width, label=label_x)
        ax.bar(x + width / 2, values_x2, width, label=label_x2)
        ax.set_xticks(x, values_y, rotation=90)
        ax.tick_params(axis="both", labelsize=16)
        ax.set_title(title, fontweight='normal', fontsize=20)
        ax.legend(loc="upper left", fontsize=14)


class PdfConverter:
    """
    Класс для конвертирования данных статистики в pdf-файл
    Attributes:
        graph_name: Название файла с графиком (путь относительный в той же папке, что и код)
        :type (str)
        excel_file_name: Название файла с excel таблицой (путь относительный в той же папке, что и код)
        :type (str)
        profession: Название профессии
        :type (str)
    """
    def __init__(self, graph_name : str, excel_file_name : str, profession : str):
        """
        @param graph_name: Название файла с графиком (путь относительный в той же папке, что и код)
        :type (str)
        @param excel_file_name: Название файла с excel таблицой (путь относительный в той же папке, что и код)
        :type (str)
        @param profession: Название профессии
        :type (str)
        """
        self.graph = graph_name
        self.excel_file = excel_file_name
        self.prof = profession

    def generate_pdf(self) -> None:
        """
        Генерирует pdf-файл из представленных данных
        @return: None
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        graph_path = os.path.abspath(self.graph)
        out_stream = xlsx2html(self.excel_file, sheet="Статистика по годам")
        out_stream.seek(0)
        pdf_template = template.render({"prof" : self.prof,
                                        "graph": graph_path,
                                        "first_table" : out_stream.read()})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class Statistic:
    """
    Класс для обработки, иницилизации данных  представления статистики
    Attributes:
        input_data: Данные представленные пользователем (Запрос, имя файла, название нужной профессии)
        :type (List[str])
    """
    def __init__(self, profession : str):
        """
        Иницилизация данных
        @param profession: Название профессии
        :type (str)
        """
        self.profession = profession

    def process_data(self, file_name : str) -> (Dict[int, int], Dict[int, int], Dict[int, int], Dict[int, int]):
        """
        Обработка данных по имени csv-файла
        @param file_name: Название файла
        :type (str)
        @return: Статистика по заданному файлу
        :type (Dict[str, str], Dict[str, str], Dict[str, str], Dict[str, str])
        """
        data = DataSet(file_name).vacancies_objects
        data_profession = [d for d in data if self.profession in d.name]
        year_salary = self.convert_to_param_salary(data)
        professions_year_salary = self.__add_missing_years(self.convert_to_param_salary(data_profession), year_salary)
        year_salary, year_vacancy = self.__convert_from_param_salary_to_dict(year_salary)
        professions_year_salary, professions_year_vacancies = self.__convert_from_param_salary_to_dict(professions_year_salary)
        return year_salary, year_vacancy, professions_year_salary, professions_year_vacancies

    def convert_to_param_salary(self, vacancies: List[Vacancy]) -> (List[YearSalary]):
        """
        Конвертирует список вакансий по параметру сравнения в список класса ParamSalary
        @param vacancies: Набор вакансий
        :type (str)
        @return: Список данных класса ParamSalary
        :type (List[ParamSalary])
        """
        param_salary = {}
        for vacancy in vacancies:
            if not param_salary.__contains__(vacancy.year):
                param_salary[vacancy.year] = YearSalary(vacancy.year, vacancy.salary)
            else:
                param_salary[vacancy.year] = param_salary[vacancy.year].add_salary(vacancy.salary)
        return [param_salary[d] for d in param_salary]

    def __convert_from_param_salary_to_dict(self, param_salary: List[YearSalary]) -> (Dict[int, int], Dict[int, int]):
        """
        Нужен для обработки списка данных класса ParamSalary и возвращения 2 словарей 1 - выбранный параметр: средняя зарплата 2 - выбранный параметр: количество вакансий
        @param param_salary: Список данных класса ParamSalary
        :type (List[ParamSalary])
        @return: Возвращает 2 словаря 1 - выбранный параметр: средняя зарплата 2 - выбранный параметр: количество вакансий
        :type (Dict[int, int], Dict[int, int])
        """
        return {x: y for x, y in zip([int(r.param) for r in param_salary],
                                     [0 if v.count_vacancies == 0 else int(v.salary / v.count_vacancies) for v in param_salary])},\
               {x: y for x, y in zip([int(r.param) for r in param_salary], [v.count_vacancies for v in param_salary])}

    def __add_missing_years(self, param_salary: List[YearSalary], year_salary : List[YearSalary]) -> List[YearSalary]:
        """
        Добавляет года, пропущенные при выборке данных
        @param param_salary:  Список данных класса ParamSalary
        :type (List[ParamSalary])
        @param year_salary: Список данных класса ParamSalary, где param - year
        :type (List[ParamSalary])
        @return:
        :type (List[ParamSalary])
        """
        years = [i.param for i in year_salary]
        s_years = [el.param for el in param_salary]
        for y in years:
            if y not in s_years:
                param_salary.insert(int(y) - int(years[0]), YearSalary(y, Salary("0", "0", "RUR", "2003-10-07T00:00:00+0400")))
                param_salary[int(y) - int(years[0])].count_vacancies = 0
        return param_salary


class CreateStatisticFiles:
    """
    Класс для создания итоговых файлов
    Attributes:
        year_salary: Набор зарплат по годам
        :type (Dict[int, int])
        year_vacancy: Набор количеств вакансий за год
        :type (Dict[int, int])
        professions_year_salary:  Набор зарплат по годам для выбранной профессии
        :type (Dict[int, int])
        professions_year_vacancies: Набор количеств вакансий за год для выбранной професии
        :type (Dict[int, int])
        profession: Профессия
        :type (str)
    """
    def __init__(self, year_salary : Dict[int, int], year_vacancy : Dict[int, int],professions_year_salary : Dict[int, int],
                 professions_year_vacancies : Dict[int, int], profession : str):
        """
        @param year_salary: Набор зарплат по годам
        :type (Dict[int, int])
        @param year_vacancy: Набор количеств вакансий за год
        :type (Dict[int, int])
        @param professions_year_salary:  Набор зарплат по годам для выбранной профессии
        :type (Dict[int, int])
        @param professions_year_vacancies: Набор количеств вакансий за год для выбранной професии
        :type (Dict[int, int])
        @param profession: Профессия
        :type (str)
        """
        self.year_salary = year_salary
        self.year_vacancy = year_vacancy
        self.professions_year_salary = professions_year_salary
        self.professions_year_vacancies = professions_year_vacancies
        self.profession = profession

    def create_files(self) -> None:
        """
        Создает файли стастики
        @return: None
        """
        output_data = {"Динамика уровня зарплат по годам:": self.year_salary,
                       "Динамика количества вакансий по годам:": self.year_vacancy,
                       "Динамика уровня зарплат по годам для выбранной профессии:": self.professions_year_salary,
                       "Динамика количества вакансий по годам для выбранной профессии:": self.professions_year_vacancies}
        [print(i, output_data[i]) for i in output_data]
        excel_file = "report.xlsx"
        report = Report(profession=self.profession,
                        years=[i for i in self.year_salary],
                        average_salary=[self.year_salary[i] for i in self.year_salary],
                        average_salary_profession=[self.professions_year_salary[i] for i in self.professions_year_salary],
                        count_vacancies_by_year=[self.year_vacancy[i] for i in self.year_vacancy],
                        count_vacancies_by_year_prof=[self.professions_year_vacancies[i] for i in self.professions_year_vacancies],
                        file_name=excel_file)
        report.generate_excel()
        graph_name = "graph.png"
        graph = Graphic(profession=self.profession,
                        years=[i for i in self.year_salary],
                        average_salary=[self.year_salary[i] for i in self.year_salary],
                        average_salary_profession=[self.professions_year_salary[i] for i in self.professions_year_salary],
                        count_vacancies_by_year=[self.year_vacancy[i] for i in self.year_vacancy],
                        count_vacancies_by_year_prof=[self.professions_year_vacancies[i] for i in self.professions_year_vacancies],
                        file_name=graph_name)
        pdf = PdfConverter(graph_name=graph_name,
                           excel_file_name=excel_file,
                           profession=self.profession)
        pdf.generate_pdf()


class InputConnect:
    """
    Класс для обработки и иницилизации данных
    Attributes:
        input_data: Данные представленные пользователем (Запрос, имя файла, название нужной профессии)
        :type (List[str])
    """
    def __init__(self):
        """
        Иницилизация данных
        """
        input_data = []
        for question in ["Введите название csv-файла: ", "Введите название профессии: "]:
            print(question, end="")
            input_data.append(input())
        self.csv_file = input_data[0]
        self.profession = input_data[1]


class GetValutesValues:
    def __init__(self, valutes):
        self.valutes = valutes

    def get_valutes(self, date):
        session = requests.Session()
        retry = Retry(connect=3, backoff_factor=0.5)
        adapter = HTTPAdapter(max_retries=retry)
        session.mount('http://', adapter)
        session.mount('https://', adapter)
        url = f"https://www.cbr.ru/scripts/XML_daily.asp?date_req=01/{date}d=1"
        res = session.get(url)
        cur_df = pd.read_xml(res.text)
        values = []
        for valute in self.valutes:
            if valute in cur_df["CharCode"].values:
                values.append(round(float(cur_df.loc[cur_df["CharCode"] == valute]["Value"].values[0].replace(',', ".")) / float(cur_df.loc[cur_df["CharCode"] == valute]["Nominal"]), 4))
            else:
                values.append(0)
        return [date] + values

    @staticmethod
    def get_date(first_date, second_date):
        res = []
        for year in range(int(first_date[:4]), int(second_date[:4]) + 1):
            num = 1
            if str(year) == first_date[:4]:
                num = int(first_date[-2:])
            for month in range(num, 13):
                if len(str(month)) == 2:
                    res.append(f"{month}/{year}")
                else:
                    res.append(f"0{month}/{year}")
                if str(year) == second_date[:4] and (str(month) == second_date[-2:] or f"0{month}" == second_date[-2:]):
                    break
        return res


directory = 'vacancies_by_year'
if __name__ == "__main__":
    year_salary, year_vacancy, professions_year_salary, professions_year_vacancies = {}, {}, {}, {}
    inp = InputConnect()
    spl = SplitCsvFileByYear(inp.csv_file, directory)
    start = time.time()
    files = [str(file) for file in pathlib.Path(f"./{directory}").iterdir()]
    stats = Statistic(inp.profession)
    with concurrent.futures.ProcessPoolExecutor() as executor:
        r = list(executor.map(stats.process_data, files))
        for el in r:
            for i, value in zip(range(4), [year_salary, year_vacancy, professions_year_salary, professions_year_vacancies]):
                value.update(el[i])
    CreateStatisticFiles(year_salary, year_vacancy, professions_year_salary, professions_year_vacancies, inp.profession).create_files()
